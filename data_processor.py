
# data_processor.py
# Handles reading the single uploaded file, fetching transcripts from URL,
# and filtering llm_extracted_json to the 12 target parameters.
# Supports new file_id column (unique) alongside mcat_id.

import io
import json
import re
import ast
import pandas as pd
import requests
from definitions import PARAMETER_ORDER

# ──────────────────────────────────────────────────────────────────────────────
# Constants
# ──────────────────────────────────────────────────────────────────────────────
FETCH_TIMEOUT = 20

# Required columns (case-insensitive alias resolution)
REQUIRED_COLUMNS = {
    "mcat_id": ["mcat_id", "mcat id", "mcatid"],
    "TranscriptionURL": [
        "transcriptionurl", "transcription_url", "transcription url",
        "transcript_url", "transcripturl", "url",
    ],
    "llm_extracted_json": [
        "llm_extracted_json", "llm extracted json", "extracted json",
        "extractedjson", "llm_json", "extracted_json", "json",
    ],
}

# Optional column — falls back to mcat_id if not found
OPTIONAL_COLUMNS = {
    "file_id": [
        "file_id", "file id", "fileid", "id", "unique_id", "uniqueid",
    ],
}

IGNORED_COLUMNS = ["pns link", "pns_link", "pnslink"]


# ──────────────────────────────────────────────────────────────────────────────
# File Reading
# ──────────────────────────────────────────────────────────────────────────────

def read_uploaded_file(uploaded_file) -> pd.DataFrame:
    """
    Read CSV or Excel files with ZERO pandas index/column inference.

    - .xlsx  → openpyxl reads every cell directly, first column always preserved
    - .csv   → csv.reader with utf-8-sig (handles Excel BOM), no index inference
    - .xls   → pandas fallback with index_col=False
    """
    import csv as _csv
    try:
        name = uploaded_file.name.lower()

        # ── CSV ───────────────────────────────────────────────────────────────
        if name.endswith(".csv"):
            raw = uploaded_file.read()
            # utf-8-sig strips the BOM that Excel adds when saving as CSV
            text = raw.decode("utf-8-sig", errors="replace")
            reader = _csv.reader(io.StringIO(text))
            all_rows = list(reader)
            if not all_rows:
                raise ValueError("CSV file is empty.")
            headers = [str(h).strip() for h in all_rows[0]]
            data_rows = all_rows[1:]

        # ── Modern Excel (.xlsx) ──────────────────────────────────────────────
        elif name.endswith(".xlsx"):
            from openpyxl import load_workbook
            raw = uploaded_file.read()
            # DO NOT use read_only=True — it skips columns that are mostly empty!
            wb = load_workbook(io.BytesIO(raw), data_only=True)
            ws = wb.active
            # values_only=True returns plain values (not Cell objects).
            # Normal mode (not read_only) reads ALL cells in the defined range,
            # including empty ones — this is critical for the file_id column
            # which has many blank cells.
            all_rows = [list(row) for row in ws.iter_rows(values_only=True)]
            wb.close()
            if not all_rows:
                raise ValueError("Excel file is empty.")
            headers = [
                str(h).strip() if h is not None else f"Unnamed_{i}"
                for i, h in enumerate(all_rows[0])
            ]
            data_rows = [
                [str(v).strip() if v is not None else "" for v in row]
                for row in all_rows[1:]
            ]

        # ── Legacy Excel (.xls) ───────────────────────────────────────────────
        elif name.endswith(".xls"):
            raw = uploaded_file.read()
            df = pd.read_excel(io.BytesIO(raw), index_col=False,
                               keep_default_na=False, dtype=str)
            df.columns = [str(c).strip() for c in df.columns]
            return df

        else:
            raise ValueError(f"Unsupported file type '{uploaded_file.name}'. Use CSV or Excel.")

        # ── Build DataFrame from raw rows ─────────────────────────────────────
        n = len(headers)
        # Pad/trim every row to the same width as the header
        data_rows = [row[:n] + [""] * max(0, n - len(row)) for row in data_rows]
        df = pd.DataFrame(data_rows, columns=headers)
        return df

    except Exception as e:
        raise RuntimeError(f"Failed to read '{uploaded_file.name}': {e}")




def _detect_column(df: pd.DataFrame, canonical: str, aliases: list[str]) -> str | None:
    """
    Find the actual column in df matching any alias (case-insensitive).
    Returns the real column name, or None if not found.
    """
    lower_map = {col.lower(): col for col in df.columns}
    for alias in aliases:
        if alias.lower() in lower_map:
            return lower_map[alias.lower()]
    return None


def load_single_file(source, file_id_col_override: str | None = None) -> pd.DataFrame:
    """
    Load the unified input file.

    Args:
        source: Either a Streamlit UploadedFile OR a pre-read pd.DataFrame
                (pass the DataFrame to avoid reading the file twice).
        file_id_col_override: If provided, use this exact column name as file_id —
                              bypasses all auto-detection logic entirely.

    Returns a DataFrame with columns: file_id | mcat_id | TranscriptionURL | llm_extracted_json
    """
    if isinstance(source, pd.DataFrame):
        df = source.copy()
    else:
        df = read_uploaded_file(source)

    # ── Resolve required columns ──────────────────────────────────────────────
    rename_map = {}
    for canonical, aliases in REQUIRED_COLUMNS.items():
        actual = _detect_column(df, canonical, aliases)
        if actual is None:
            raise KeyError(
                f"Required column '{canonical}' not found. "
                f"Tried: {aliases}. Available: {list(df.columns)}"
            )
        rename_map[actual] = canonical

    # ── Resolve file_id column ────────────────────────────────────────────────
    # claimed_actual_cols: original column names already assigned to required fields
    claimed_actual_cols = set(rename_map.keys())

    # is_required_override: True when the user picked a required column (e.g. mcat_id)
    is_required_override = file_id_col_override in claimed_actual_cols
    canonical_of_override = rename_map.get(file_id_col_override or "")  # e.g. "mcat_id"

    if file_id_col_override and file_id_col_override in df.columns:
        file_id_col = file_id_col_override  # could be required or not
    else:
        # Auto-detect from remaining (unclaimed) columns
        remaining_cols = [c for c in df.columns if c not in claimed_actual_cols]
        lower_remaining = {c.lower(): c for c in remaining_cols}
        file_id_col = None

        # Pass 1: exact alias match (also catches Unnamed_0)
        for alias in ["file_id", "file id", "fileid", "file_Id", "FileID", "File_ID",
                      "unique_id", "uniqueid", "uid", "unnamed_0"]:
            if alias.lower() in lower_remaining:
                file_id_col = lower_remaining[alias.lower()]
                break

        # Pass 2: any col containing both 'file' and 'id'
        if not file_id_col:
            for c in remaining_cols:
                cl = c.lower().replace(" ", "").replace("_", "")
                if "file" in cl and "id" in cl:
                    file_id_col = c
                    break

        # Pass 3: any remaining col ending with '_id' or ' id'
        if not file_id_col:
            for c in remaining_cols:
                if c.lower().endswith("_id") or c.lower().endswith(" id"):
                    file_id_col = c
                    break

    df = df.rename(columns=rename_map)  # apply required-column renames

    if is_required_override and canonical_of_override:
        # User picked a required column (e.g. mcat_id) as file_id — COPY, never rename,
        # so the original canonical column stays intact in the DataFrame.
        df["file_id"] = df[canonical_of_override].copy()
    elif file_id_col and file_id_col in df.columns and file_id_col != "file_id":
        df = df.rename(columns={file_id_col: "file_id"})
    elif file_id_col == "file_id" or "file_id" in df.columns:
        pass  # already correctly named
    else:
        df["file_id"] = df["mcat_id"].copy()  # last-resort fallback

    # ── Drop ignored columns ──────────────────────────────────────────────────
    keep = {"file_id", "mcat_id", "TranscriptionURL", "llm_extracted_json"}
    cols_to_drop = [c for c in df.columns if c.lower() in IGNORED_COLUMNS and c not in keep]
    df = df.drop(columns=cols_to_drop, errors="ignore")

    # ── Keep only target columns ──────────────────────────────────────────────
    for col in ["file_id", "mcat_id", "TranscriptionURL", "llm_extracted_json"]:
        if col not in df.columns:
            df[col] = df.get("mcat_id", "")

    df = df[["file_id", "mcat_id", "TranscriptionURL", "llm_extracted_json"]].copy()

    # ── Normalize & fill blanks ──────────────────────────────────────────────────
    df["file_id"]  = df["file_id"].astype(str).str.strip()
    df["mcat_id"]  = df["mcat_id"].astype(str).str.strip()
    df["TranscriptionURL"] = df["TranscriptionURL"].astype(str).str.strip()

    df = df.dropna(subset=["mcat_id", "TranscriptionURL"])
    df = df[df["mcat_id"].str.lower().isin(["nan","none",""]) == False]  # noqa

    # KEY FIX: If file_id cell is blank/NaN for a row, use that row’s mcat_id
    # instead of dropping the row. Dropping caused the fallback where both columns
    # showed the same value.
    _NULL_VALS = {"", "nan", "none", "null", "na", "n/a"}
    empty_mask = df["file_id"].str.lower().isin(_NULL_VALS)
    df.loc[empty_mask, "file_id"] = df.loc[empty_mask, "mcat_id"]

    df = df.drop_duplicates(subset="file_id", keep="first")
    df = df.reset_index(drop=True)

    return df


# ──────────────────────────────────────────────────────────────────────────────
# Transcript URL Fetching
# ──────────────────────────────────────────────────────────────────────────────

def fetch_transcript_from_url(url: str) -> str:
    """
    Fetch raw transcript text from TranscriptionURL.
    Handles plain-text and JSON responses.
    """
    if not url or url.lower() in ("nan", "none", ""):
        raise ValueError("TranscriptionURL is empty or missing.")

    try:
        response = requests.get(url, timeout=FETCH_TIMEOUT)
        response.raise_for_status()
    except requests.exceptions.Timeout:
        raise RuntimeError(f"Timeout fetching transcript: {url}")
    except requests.exceptions.ConnectionError:
        raise RuntimeError(f"Connection error fetching transcript: {url}")
    except requests.exceptions.HTTPError as e:
        raise RuntimeError(f"HTTP {response.status_code} error: {url} → {e}")
    except requests.exceptions.RequestException as e:
        raise RuntimeError(f"Request failed for {url}: {e}")

    content_type = response.headers.get("Content-Type", "").lower()

    if "json" in content_type:
        try:
            data = response.json()
            for key in ("transcript", "text", "content", "transcription", "body"):
                if key in data and isinstance(data[key], str):
                    return data[key].strip()
            return json.dumps(data, ensure_ascii=False)
        except json.JSONDecodeError:
            pass

    text = response.text.strip()
    if not text:
        raise RuntimeError(f"Empty response body from: {url}")
    return text


# ──────────────────────────────────────────────────────────────────────────────
# JSON Parsing & Filtering
# ──────────────────────────────────────────────────────────────────────────────

# Explicit path mapping for the known nested JSON structure
_EXPLICIT_PATHS: dict[str, list[tuple[list[str], int | None]]] = {
    "call_type":          [(["metadata", "call_type"],          None)],
    "call_purpose":       [(["metadata", "call_purpose"],        None)],
    "all_languages":      [(["metadata", "all_languages"],       None)],
    "primary_language":   [(["metadata", "primary_language"],    None)],
    "price":              [(["products", "price"],               0)],
    "in_stock":           [(["products", "in_stock"],            0)],
    "product_name":       [(["products", "product_name"],        0)],
    "specifications":     [(["products", "specifications"],      0)],
    "quantity_required":  [(["products", "quantity_required"],   0)],
    "is_buyer_interested":[(["products", "is_buyer_interested"], 0)],
    "buyer_next_steps":   [(["next_steps", "buyer_next_steps"],  None)],
    "seller_next_steps":  [(["next_steps", "seller_next_steps"], None)],
}


def _norm_key(key: str) -> str:
    """Lowercase + strip spaces/underscores/hyphens."""
    return re.sub(r'[\s_\-]', '', key.lower())


def _extract_explicit(parsed: dict, param: str):
    """Follow explicit path for a parameter. Returns (value, found: bool)."""
    paths = _EXPLICIT_PATHS.get(param, [])
    for path_keys, list_index in paths:
        node = parsed
        ok = True
        for i, key in enumerate(path_keys):
            if not isinstance(node, dict):
                ok = False; break
            if key in node:
                node = node[key]
            else:
                lower_map = {k.lower(): k for k in node}
                if key.lower() in lower_map:
                    node = node[lower_map[key.lower()]]
                else:
                    ok = False; break
            if list_index is not None and i == 0 and isinstance(node, list):
                node = node[list_index] if len(node) > list_index else None
                if node is None:
                    ok = False; break
        if ok:
            return node, True
    return None, False


def _recursive_find(obj, target_norm_keys: dict[str, str]) -> dict:
    """Recursively search for key matches anywhere in obj."""
    found: dict[str, object] = {}

    def _walk(node):
        if isinstance(node, dict):
            for k, v in node.items():
                nk = _norm_key(str(k))
                if nk in target_norm_keys and target_norm_keys[nk] not in found:
                    found[target_norm_keys[nk]] = v
                _walk(v)
        elif isinstance(node, list):
            for item in node:
                _walk(item)

    _walk(obj)
    return found


def _try_parse_raw(raw_json_str) -> tuple[dict | None, str | None]:
    """Parse raw string to dict."""
    raw = str(raw_json_str).strip()
    if raw.lower() in ("nan", "none", "null", ""):
        return None, "Empty JSON value."
    try:
        parsed = json.loads(raw)
        if isinstance(parsed, dict):
            return parsed, None
    except json.JSONDecodeError:
        pass
    try:
        parsed = ast.literal_eval(raw)
        if isinstance(parsed, dict):
            return parsed, None
    except Exception:
        pass
    return None, "Could not parse the JSON string."


def parse_and_filter_json(raw_json_str) -> dict:
    """
    Parse llm_extracted_json and extract only the 12 target parameters.
    Strategy 1: explicit path mapping.
    Strategy 2: recursive key search fallback.
    """
    if not raw_json_str or isinstance(raw_json_str, float):
        return {}
    if str(raw_json_str).strip().lower() in ("nan", "none", "null", ""):
        return {}

    parsed, err = _try_parse_raw(raw_json_str)
    if parsed is None:
        return {"_parse_error": err or "Parse failed."}

    filtered: dict[str, object] = {}

    for param in PARAMETER_ORDER:
        value, found = _extract_explicit(parsed, param)
        if found:
            filtered[param] = value

    missing = [p for p in PARAMETER_ORDER if p not in filtered]
    if missing:
        norm_targets: dict[str, str] = {_norm_key(p): p for p in missing}
        for k, v in _recursive_find(parsed, norm_targets).items():
            if k not in filtered:
                filtered[k] = v

    return filtered


def get_raw_json_keys(raw_json_str) -> tuple[list[str], str | None]:
    """Return leaf-level paths for the JSON Key Inspector UI."""
    if not raw_json_str or isinstance(raw_json_str, float):
        return [], "llm_extracted_json cell is empty or NaN."
    parsed, err = _try_parse_raw(raw_json_str)
    if parsed is None:
        return [], err or "Parse failed."

    def _collect(node, prefix="") -> list[str]:
        paths = []
        if isinstance(node, dict):
            for k, v in node.items():
                full = f"{prefix}.{k}" if prefix else k
                paths.extend(_collect(v, full) if isinstance(v, (dict, list)) else [full])
        elif isinstance(node, list):
            for i, item in enumerate(node):
                paths.extend(_collect(item, f"{prefix}[{i}]"))
        else:
            paths.append(prefix)
        return paths

    return _collect(parsed), None


# ──────────────────────────────────────────────────────────────────────────────
# Accessors
# ──────────────────────────────────────────────────────────────────────────────

def get_all_file_ids(df: pd.DataFrame) -> list[str]:
    """Return sorted list of all unique file_ids."""
    return sorted(df["file_id"].tolist())


def get_all_mcat_ids(df: pd.DataFrame) -> list[str]:
    """Return sorted list of all mcat_ids (kept for backwards compat)."""
    return sorted(df["mcat_id"].tolist())


def get_row_by_file_id(df: pd.DataFrame, file_id: str) -> dict:
    """
    Retrieve data for a file_id, fetching transcript live from URL.

    Returns:
        {file_id, mcat_id, transcript, filtered_json, transcript_url}
    """
    rows = df[df["file_id"] == file_id]
    if rows.empty:
        raise KeyError(f"file_id '{file_id}' not found in loaded data.")
    row = rows.iloc[0]
    url = str(row["TranscriptionURL"])
    transcript = fetch_transcript_from_url(url)
    filtered = parse_and_filter_json(row["llm_extracted_json"])
    return {
        "file_id": file_id,
        "mcat_id": str(row["mcat_id"]),
        "transcript": transcript,
        "filtered_json": filtered,
        "transcript_url": url,
    }


# Keep old function name for any lingering references
def get_row_by_mcat_id(df: pd.DataFrame, mcat_id: str) -> dict:
    """Alias: find first row matching mcat_id and use its file_id."""
    rows = df[df["mcat_id"] == mcat_id]
    if rows.empty:
        raise KeyError(f"mcat_id '{mcat_id}' not found.")
    return get_row_by_file_id(df, str(rows.iloc[0]["file_id"]))
